from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from aerosandbox import XFoil, Airfoil
import numpy as np

def main():
    while True:
        airfoil = input("NACA")
        if len(airfoil) == 4:
            break
        print("Invalid Airfoil: Type NACA 4-Digit")

    while True:
        try:
            Re = int(input("Re = "))
            if Re < 0:
                raise ValueError
            break
        except ValueError:
            print("Invalid Reynolds Number")
            continue
        
        

    xf = XFoil(
        airfoil=Airfoil(f'naca{airfoil}').repanel(n_points_per_side=200),
        Re = Re
        )

    data = xf.alpha(np.linspace(-2, 2, 5))
    data_keys =[]
    for props in data:
        data_keys.append(props)

    wb = Workbook()
    ws = wb.active

    for col in range(len(data_keys)):
        char = get_column_letter(col + 1)
        ws[f'{char}1'] = data_keys[col]
        for row in range(len(data['CL'])):
            ws[f'{char}{row + 2}'] = np.ndarray.tolist(data[data_keys[col]])[row]
        
    wb.save(fr"C:\Users\reach\Desktop\2D_Airfoil_Easy_Anaysis_Project\Quick-and-Easy-2D-Airfoil-Analysis\Excel_Data\{airfoil}_Re{round(Re)}.xlsx")

if __name__ == '__main__':
    main()